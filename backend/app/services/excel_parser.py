import hashlib
import mimetypes
import os
import zipfile
from difflib import SequenceMatcher
from typing import Optional

from openpyxl import load_workbook
from pydantic import ValidationError

from app.schemas.presensi import (
    EXCEL_HEADER_COLS,
    PRESENSI_COUNTER_COLS,
    ParseError,
    ParseMetadata,
    ParseResult,
    ParseWarning,
    PresensiRowData,
)


class ExcelPresensiParser:
    MAX_SIZE_MB = 10
    WARNING_COUNTER_THRESHOLD = 20
    WARNING_RARE_COUNTER_THRESHOLD = 5
    RARE_COUNTERS = {"CKAP", "CB", "CM", "CT", "CS"}
    EXPECTED_MIME = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    OPD_FUZZY_THRESHOLD = 0.8
    MAX_DAYS_IN_MONTH = 31

    def __init__(self, opd_names: Optional[set[str]] = None):
        self.opd_names = opd_names or set()

    def parse(self, file_path: str) -> ParseResult:
        self._validate_file(file_path)
        file_hash = self._compute_hash(file_path)
        file_size = os.path.getsize(file_path)
        file_name = os.path.basename(file_path)

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        try:
            self._validate_headers(headers)
        except ValueError as e:
            wb.close()
            return ParseResult(
                errors=[ParseError(row_number=0, column=None, value=None, reason=str(e))],
                metadata=ParseMetadata(
                    file_hash=file_hash,
                    file_size_bytes=file_size,
                    file_name=file_name,
                    total_rows=0,
                    success_rows=0,
                    error_rows=0,
                    warning_count=0,
                ),
            )

        rows: list[PresensiRowData] = []
        errors: list[ParseError] = []
        warnings: list[ParseWarning] = []
        total_rows = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None for v in row):
                continue
            total_rows += 1
            try:
                data = dict(zip(EXCEL_HEADER_COLS, row, strict=False))
                resolved = self._resolve_row(data, row_idx, warnings)
                parsed = PresensiRowData(**resolved)
                rows.append(parsed)
            except ValidationError as e:
                for err in e.errors():
                    loc = str(err["loc"][0]) if err.get("loc") else None
                    errors.append(ParseError(
                        row_number=row_idx,
                        column=loc,
                        value=err.get("input"),
                        reason=err["msg"],
                    ))

        wb.close()

        return ParseResult(
            rows=rows,
            errors=errors,
            warnings=warnings,
            metadata=ParseMetadata(
                file_hash=file_hash,
                file_size_bytes=file_size,
                file_name=file_name,
                total_rows=total_rows,
                success_rows=len(rows),
                error_rows=len(errors),
                warning_count=len(warnings),
            ),
        )

    def _validate_file(self, file_path: str) -> None:
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        ext = os.path.splitext(file_path)[1].lower()
        if ext != ".xlsx":
            raise ValueError(f"Only .xlsx files are accepted, got: {ext}")

        mime, _ = mimetypes.guess_type(file_path)
        if mime and mime != self.EXPECTED_MIME:
            raise ValueError(f"Invalid MIME type: {mime}. Expected: {self.EXPECTED_MIME}")

        file_size = os.path.getsize(file_path)
        if file_size > self.MAX_SIZE_MB * 1024 * 1024:
            raise ValueError(
                f"File size {file_size / 1024 / 1024:.1f}MB exceeds "
                f"maximum {self.MAX_SIZE_MB}MB"
            )

        if not zipfile.is_zipfile(file_path):
            raise ValueError("File is not a valid .xlsx (not a ZIP archive)")

    @staticmethod
    def _compute_hash(file_path: str) -> str:
        sha256 = hashlib.sha256()
        with open(file_path, "rb") as f:
            while chunk := f.read(8192):
                sha256.update(chunk)
        return sha256.hexdigest()

    HEADER_ALIASES = {
        "NO": "NO URUT",
        "UNIT KERJA ": "UNIT KERJA",
        "PEGAWAI": "NAMA",
        "NAMA PEGAWAI": "NAMA",
        "OPD": "UNIT KERJA",
        "INSTANSI": "UNIT KERJA",
    }

    def _validate_headers(self, headers: list[str | None]) -> list[str]:
        cleaned = [str(h).strip() if h is not None else "" for h in headers]
        
        # Check if this is a single OPD detail template (e.g. has "PEGAWAI" or "TOTAL PENGURANGAN (%)")
        is_detail_template = any(h.upper() in ("PEGAWAI", "TOTAL PENGURANGAN (%)") for h in cleaned if h)
        
        if is_detail_template:
            normalized = []
            for h in cleaned:
                h_upper = h.upper()
                if h_upper in ("NO", "NIP"):
                    mapped = "NIP"
                elif h_upper in ("PEGAWAI", "NAMA"):
                    mapped = "NAMA"
                else:
                    mapped = self.HEADER_ALIASES.get(h_upper, h)
                normalized.append(mapped)
            return normalized

        # Standard master file validation
        cleaned_mapped = [self.HEADER_ALIASES.get(h, h) for h in cleaned]
        missing = set(EXCEL_HEADER_COLS) - set(cleaned_mapped)
        if missing:
            raise ValueError(
                f"Missing required columns: {sorted(missing)}. "
                f"Expected {len(EXCEL_HEADER_COLS)} columns."
            )
        extra = set(cleaned_mapped) - set(EXCEL_HEADER_COLS)
        if extra:
            raise ValueError(f"Unexpected extra columns: {sorted(extra)}")

        return cleaned_mapped

    def parse(self, file_path: str, default_tahun: int = 2026, default_bulan: int = 6, default_unit: str = "BKPSDM") -> ParseResult:
        self._validate_file(file_path)
        file_hash = self._compute_hash(file_path)
        file_size = os.path.getsize(file_path)
        file_name = os.path.basename(file_path)

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        try:
            normalized_headers = self._validate_headers(headers)
        except ValueError as e:
            wb.close()
            return ParseResult(
                errors=[ParseError(row_number=0, column=None, value=None, reason=str(e))],
                metadata=ParseMetadata(
                    file_hash=file_hash,
                    file_size_bytes=file_size,
                    file_name=file_name,
                    total_rows=0,
                    success_rows=0,
                    error_rows=0,
                    warning_count=0,
                ),
            )

        rows: list[PresensiRowData] = []
        errors: list[ParseError] = []
        warnings: list[ParseWarning] = []
        total_rows = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None for v in row):
                continue
            
            # Map row by header name
            raw = {hdr: val for hdr, val in zip(normalized_headers, row) if hdr}
            
            # Ignore summary rows at footer
            nip_val = str(raw.get("NIP", "")).strip()
            nama_val = str(raw.get("NAMA", "")).strip()
            if nip_val.upper() in ("TOTAL", "JUMLAH") or nama_val.upper().startswith("JLH") or nama_val.upper().startswith("TOTAL"):
                continue

            total_rows += 1
            try:
                resolved = self._resolve_row(raw, row_idx, warnings, default_tahun, default_bulan, default_unit)
                parsed = PresensiRowData(**resolved)
                rows.append(parsed)
            except ValidationError as e:
                for err in e.errors():
                    loc = str(err["loc"][0]) if err.get("loc") else None
                    errors.append(ParseError(
                        row_number=row_idx,
                        column=loc,
                        value=err.get("input"),
                        reason=err["msg"],
                    ))

        wb.close()

        return ParseResult(
            rows=rows,
            errors=errors,
            warnings=warnings,
            metadata=ParseMetadata(
                file_hash=file_hash,
                file_size_bytes=file_size,
                file_name=file_name,
                total_rows=total_rows,
                success_rows=len(rows),
                error_rows=len(errors),
                warning_count=len(warnings),
            ),
        )

    COUNTER_COLS_LOWER = {c.lower() for c in PRESENSI_COUNTER_COLS}

    def _resolve_row(
        self,
        raw: dict[str, any],
        row_idx: int,
        warnings: list[ParseWarning],
        default_tahun: int = 2026,
        default_bulan: int = 6,
        default_unit: str = "BKPSDM",
    ) -> dict[str, any]:
        resolved: dict[str, any] = {}

        # Resolve primary columns
        no_urut = raw.get("NO URUT")
        if not no_urut or not str(no_urut).isdigit():
            no_urut = row_idx - 1
        resolved["NO URUT"] = int(no_urut)

        nip = raw.get("NIP")
        if isinstance(nip, str):
            nip = nip.strip().lstrip("'")
        elif isinstance(nip, (int, float)):
            nip = str(int(nip))
        resolved["NIP"] = nip

        nama = raw.get("NAMA")
        if not nama and nip:
            nama = f"Pegawai {nip}"
        resolved["NAMA"] = str(nama).strip() if nama else "Unknown"

        tahun = raw.get("TAHUN")
        if not tahun or not str(tahun).isdigit():
            tahun = default_tahun
        resolved["TAHUN"] = int(tahun)

        bulan = raw.get("BULAN")
        if not bulan or not str(bulan).isdigit():
            bulan = default_bulan
        resolved["BULAN"] = int(bulan)

        unit = raw.get("UNIT KERJA")
        if not unit:
            unit = default_unit
        resolved["UNIT KERJA"] = str(unit).strip()

        # OPD fuzzy match check
        unit_str = resolved["UNIT KERJA"]
        if self.opd_names and unit_str and unit_str not in self.opd_names:
            best_match = self._fuzzy_match_opd(unit_str)
            if best_match:
                warnings.append(ParseWarning(
                    row_number=row_idx,
                    column="UNIT KERJA",
                    value=unit_str,
                    reason=f"OPD not in master. Did you mean '{best_match}'?",
                ))
            else:
                warnings.append(ParseWarning(
                    row_number=row_idx,
                    column="UNIT KERJA",
                    value=unit_str,
                    reason="OPD not found in master",
                ))

        # Resolve counters
        counter_sum = 0
        for col in PRESENSI_COUNTER_COLS:
            val = raw.get(col)
            col_lower = col.lower()
            if val is None or val == "":
                resolved[col_lower] = 0
                continue
            try:
                val = int(float(str(val)))
            except (ValueError, TypeError):
                val = 0
            resolved[col_lower] = val
            counter_sum += val

        if counter_sum > self.MAX_DAYS_IN_MONTH:
            warnings.append(ParseWarning(
                row_number=row_idx,
                column=None,
                value=counter_sum,
                reason=f"Total counter sum ({counter_sum}) exceeds {self.MAX_DAYS_IN_MONTH} days",
            ))

        for col in self.RARE_COUNTERS:
            lower_col = col.lower()
            val = resolved.get(lower_col, 0)
            if isinstance(val, (int, float)) and val > self.WARNING_RARE_COUNTER_THRESHOLD:
                warnings.append(ParseWarning(
                    row_number=row_idx,
                    column=col,
                    value=val,
                    reason=f"Unusually high value for {col}",
                ))

        return resolved

    def _fuzzy_match_opd(self, name: str) -> Optional[str]:
        if not self.opd_names:
            return None
        best_ratio = 0.0
        best_match = None
        for opd in self.opd_names:
            ratio = SequenceMatcher(None, name.lower(), opd.lower()).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                best_match = opd
        if best_match and best_ratio >= self.OPD_FUZZY_THRESHOLD:
            return best_match
        return None
